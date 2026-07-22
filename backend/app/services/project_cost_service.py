from datetime import datetime
from decimal import Decimal
from io import BytesIO
from uuid import UUID

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload

from app.models.business_document import BusinessDocument
from app.models.project_cost import ProjectCost
from app.models.task import Attachment
from app.repositories.project_cost_repo import ProjectCostRepository
from app.repositories.task_repo import AttachmentRepository

from app.services.number_generator import generate_project_cost_no


class ProjectCostService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.repo = ProjectCostRepository(db)
        self.attachment_repo = AttachmentRepository(db)

    async def _sync_document_cost(self, document_id: UUID | None) -> None:
        """Re-calculate order cost when the associated document is an order."""
        if not document_id:
            return
        result = await self.db.execute(
            select(BusinessDocument.doc_type).where(BusinessDocument.id == document_id)
        )
        doc_type = result.scalar_one_or_none()
        if doc_type != "order":
            return
        from app.services.order_service import OrderService
        order_svc = OrderService(self.db)
        await order_svc.auto_calculate_cost(document_id)

    async def list_costs(
        self,
        page: int,
        page_size: int,
        order_id: UUID | None = None,
        quote_id: UUID | None = None,
        source_type: str | None = None,
        category: str | None = None,
        date_from: str | None = None,
        date_to: str | None = None,
    ) -> tuple[list, int]:
        skip = (page - 1) * page_size
        costs, total = await self.repo.list_costs(skip, page_size, order_id, quote_id, source_type, category, date_from, date_to)
        result = [self._to_dict(c) for c in costs]
        # Populate attachment counts
        if result:
            cost_ids = [c.id for c in costs]
            counts = await self._get_attachment_counts(cost_ids)
            for d in result:
                d["attachment_count"] = counts.get(d["id"], 0)
        return result, total

    async def get_cost(self, cost_id: UUID) -> dict | None:
        c = await self.repo.get_by_id(cost_id)
        if not c:
            return None
        d = self._to_dict(c)
        atts = await self.attachment_repo.get_by_task("project_cost", cost_id)
        d["attachments"] = [
            {
                "id": str(a.id),
                "filename": a.filename,
                "file_path": a.file_path,
                "file_size": a.file_size,
                "file_type": a.file_type,
                "category": a.category,
                "uploaded_by": str(a.uploaded_by) if a.uploaded_by else None,
                "created_at": a.created_at.isoformat() if a.created_at else None,
            }
            for a in atts
        ]
        d["attachment_count"] = len(atts)
        return d

    async def create_cost(self, data: dict, created_by: UUID) -> dict:
        # Resolve document_id from backward-compat params
        document_id = None
        if data.get("document_id"):
            document_id = UUID(data["document_id"])
        elif data.get("order_id"):
            document_id = UUID(data["order_id"])
        elif data.get("quote_id"):
            document_id = UUID(data["quote_id"])

        customer_id_val = None
        project_name_val = None
        doc_no_val = None
        doc_type_val = None

        if document_id:
            result = await self.db.execute(
                select(BusinessDocument).where(BusinessDocument.id == document_id)
            )
            doc = result.scalar_one_or_none()
            if not doc:
                raise ValueError("业务单据不存在")
            customer_id_val = doc.customer_id
            project_name_val = doc.project_name
            doc_no_val = doc.doc_no
            doc_type_val = doc.doc_type

        # Resolve document_item_id from backward-compat params
        document_item_id = None
        if data.get("document_item_id"):
            document_item_id = UUID(data["document_item_id"])
        elif data.get("order_item_id"):
            document_item_id = UUID(data["order_item_id"])
        elif data.get("quote_item_id"):
            document_item_id = UUID(data["quote_item_id"])

        debt_amount = data.get("debt_amount")
        if debt_amount is not None:
            debt_amount = float(debt_amount)
        else:
            debt_amount = 0

        cost = ProjectCost(
            cost_no=await generate_project_cost_no(self.db),
            document_id=document_id,
            customer_id=customer_id_val,
            category=data["category"],
            amount=data["amount"],
            description=data.get("description"),
            summary=data.get("summary"),
            cost_date=datetime.fromisoformat(data["cost_date"]) if data.get("cost_date") else None,
            receipt_url=data.get("receipt_url"),
            quantity=data.get("quantity"),
            specification=data.get("specification"),
            unit=data.get("unit"),
            unit_price=data.get("unit_price"),
            remark=data.get("remark"),
            document_item_id=document_item_id,
            payment_method=data.get("payment_method"),
            payee_company_name=data.get("payee_company_name"),
            debt_amount=debt_amount,
            is_debt=debt_amount > 0,
            is_settled=False,
            created_by=created_by,
        )
        # Explicitly set group_name after constructor (SQLAlchemy kwarg issue)
        if data.get("group_name"):
            cost.group_name = data["group_name"]
        await self.repo.create(cost)
        if document_id:
            await self._sync_document_cost(document_id)
        return {
            "id": str(cost.id),
            "cost_no": cost.cost_no,
            "source_type": doc_type_val,
            "document_id": str(cost.document_id) if cost.document_id else None,
            "doc_no": doc_no_val,
            "order_id": str(cost.document_id) if cost.document_id else None,
            "quote_id": str(cost.document_id) if cost.document_id and doc_type_val == "quote" else None,
            "order_no": doc_no_val if doc_type_val == "order" else None,
            "quote_no": doc_no_val if doc_type_val == "quote" else None,
            "document_item_id": str(cost.document_item_id) if cost.document_item_id else None,
            "order_item_id": str(cost.document_item_id) if cost.document_item_id and doc_type_val == "order" else None,
            "quote_item_id": str(cost.document_item_id) if cost.document_item_id and doc_type_val == "quote" else None,
            "group_name": cost.group_name,
            "order_item_name": None,  # populated via document_item relationship on read
            "quote_item_name": None,
            "customer_id": str(cost.customer_id) if cost.customer_id else None,
            "customer_name": None,  # populated by list query via relationship
            "project_name": project_name_val,
            "category": cost.category,
            "amount": float(cost.amount),
            "quantity": float(cost.quantity) if cost.quantity else None,
            "specification": cost.specification,
            "unit": cost.unit,
            "unit_price": float(cost.unit_price) if cost.unit_price else None,
            "payment_method": cost.payment_method,
            "payee_company_name": cost.payee_company_name,
            "debt_amount": float(cost.debt_amount) if cost.debt_amount else None,
            "is_debt": cost.is_debt,
            "is_settled": cost.is_settled,
            "settled_at": cost.settled_at.isoformat() if cost.settled_at else None,
            "description": cost.description,
            "summary": cost.summary,
            "cost_date": cost.cost_date.isoformat() if cost.cost_date else None,
            "receipt_url": cost.receipt_url,
            "remark": cost.remark,
            "created_by": str(cost.created_by) if cost.created_by else None,
            "created_at": cost.created_at.isoformat() if cost.created_at else None,
        }

    async def update_cost(self, cost_id: UUID, data: dict) -> dict:
        c = await self.repo.get_by_id(cost_id)
        if not c:
            raise ValueError("项目成本记录不存在")
        if "cost_date" in data and data["cost_date"] is not None:
            data = {**data, "cost_date": datetime.fromisoformat(data["cost_date"])}
        await self.repo.update(c, data)
        await self._sync_document_cost(c.document_id)
        # Re-fetch with relationships loaded for response
        c = await self.repo.get_by_id(cost_id)
        return self._to_dict(c)

    async def delete_cost(self, cost_id: UUID) -> None:
        c = await self.repo.get_by_id(cost_id)
        if not c:
            raise ValueError("项目成本记录不存在")
        document_id = c.document_id
        await self.repo.soft_delete(c)
        if document_id:
            await self._sync_document_cost(document_id)

    async def batch_delete_costs(self, cost_ids: list[UUID]) -> int:
        from app.models.project_cost import ProjectCost
        from sqlalchemy import select
        # Collect document_ids before deletion for cost sync
        result = await self.db.execute(
            select(ProjectCost.document_id)
            .where(ProjectCost.id.in_(cost_ids), ProjectCost.deleted_at.is_(None))
        )
        document_ids = {row[0] for row in result.all() if row[0]}
        deleted = await self.repo.batch_soft_delete(cost_ids)
        for did in document_ids:
            await self._sync_document_cost(did)
        return deleted

    async def get_costs_summary(self, document_ids: list[UUID]) -> dict[str, float]:
        """Return {document_id: total_cost} for a batch of documents."""
        return await self.repo.get_costs_summary(document_ids)

    async def list_debts(
        self,
        page: int,
        page_size: int,
        keyword: str | None = None,
        is_settled: bool | None = None,
    ) -> tuple[list, int]:
        """List all cost debts."""
        from sqlalchemy import or_
        from app.models.customer import Customer as CustomerModel

        skip = (page - 1) * page_size
        q = select(ProjectCost).options(
            selectinload(ProjectCost.document),
            selectinload(ProjectCost.customer),
        ).where(
            ProjectCost.deleted_at.is_(None),
            ProjectCost.is_debt == True,
        )
        if is_settled is not None:
            q = q.where(ProjectCost.is_settled == is_settled)
        if keyword:
            fuzzy = f"%{keyword}%"
            q = q.join(ProjectCost.document).join(ProjectCost.customer, isouter=True).where(
                or_(
                    BusinessDocument.doc_no.ilike(fuzzy),
                    BusinessDocument.project_name.ilike(fuzzy),
                    CustomerModel.name.ilike(fuzzy),
                )
            )

        # Count
        count_q = select(func.count()).select_from(q.subquery())
        total = (await self.db.execute(count_q)).scalar()

        q = q.order_by(ProjectCost.created_at.desc()).offset(skip).limit(page_size)
        result = await self.db.execute(q)
        costs = list(result.scalars().all())

        result_list = []
        for c in costs:
            d = self._to_dict(c)
            result_list.append(d)
        return result_list, total

    async def settle_debt(self, cost_id: UUID, settle_data: dict) -> dict:
        """Settle a cost debt (write-off)."""
        from datetime import datetime, timezone

        c = await self.repo.get_by_id(cost_id)
        if not c:
            raise ValueError("成本记录不存在")
        if not c.is_debt:
            raise ValueError("该记录不是欠款记录")
        if c.is_settled:
            raise ValueError("该欠款已结清")

        c.is_settled = True
        c.settled_at = datetime.now(timezone.utc)
        c.payment_method = settle_data.get("payment_method", c.payment_method or "转账支付")
        if settle_data.get("remark"):
            c.remark = (c.remark or "") + f" [结清: {settle_data['remark']}]"
        await self.db.flush()

        # Also create a formal cost entry for the settled debt amount
        settle_cost = ProjectCost(
            cost_no=await generate_project_cost_no(self.db),
            document_id=c.document_id,
            customer_id=c.customer_id,
            category=c.category,
            amount=settle_data.get("settle_amount", float(c.debt_amount or 0)),
            payment_method=settle_data.get("payment_method", "转账支付"),
            debt_amount=0,
            is_debt=False,
            is_settled=False,
            description=f"欠款冲红 - 原成本编号 {c.cost_no}",
            cost_date=datetime.now(timezone.utc),
            remark=settle_data.get("remark"),
            created_by=c.created_by,
        )
        await self.repo.create(settle_cost)
        await self._sync_document_cost(c.document_id)

        return self._to_dict(c)

    async def import_from_excel(self, file: BytesIO, created_by: UUID, order_id: UUID | None = None, quote_id: UUID | None = None, source_type: str = "order") -> dict:
        """Parse Excel file and create ProjectCost records.
        When order_id or quote_id is provided, all rows are assigned to that entity and the
        Excel only needs columns: 成本类别, 金额, 描述(可选), 成本日期(可选), 备注(可选).
        When neither is provided, the Excel must include column: 订单编号/报价单编号, 成本类别, 金额, ...
        """
        import openpyxl

        wb = openpyxl.load_workbook(file, read_only=True)
        ws = wb.active
        # Read header row to build column name -> index mapping
        headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        if not headers or not headers[0]:
            return {"created": 0, "errors": [{"row": 1, "error": "Excel 文件缺少表头行"}]}
        col_map = {}
        for col_idx, h in enumerate(headers[0]):
            if h is None:
                continue
            name = str(h).strip()
            col_map[name] = col_idx

        rows = list(ws.iter_rows(min_row=2, values_only=True))

        created = 0
        errors = []

        def get(col_name: str, default=None):
            """Get value from row by column name."""
            idx = col_map.get(col_name)
            if idx is None:
                return default
            return row[idx] if idx < len(row) else default

        for i, row in enumerate(rows, start=2):
            if not row:
                continue
            try:
                if order_id or quote_id:
                    # Import within document context — document_id pre-set
                    category = str(get("成本类别") or "").strip()
                    payment_method = str(get("付款方式") or "").strip() or None
                    payee_company_name = str(get("收款公司") or "").strip() or None
                    quantity_val = get("数量")
                    quantity = float(quantity_val) if quantity_val else None
                    specification = str(get("规格尺寸") or "").strip() or None
                    unit = str(get("单位") or "").strip() or None
                    unit_price_val = get("单价")
                    unit_price = float(unit_price_val) if unit_price_val else None
                    amount_val = get("金额")
                    amount = float(amount_val) if amount_val else 0
                    debt_val = get("欠款金额")
                    debt_amount = float(debt_val) if debt_val else 0
                    cost_date_str = str(get("成本日期") or "").strip() or None
                    description = str(get("说明") or "").strip() or None
                    summary = str(get("成本摘要") or "").strip() or None
                    remark = str(get("备注") or "").strip() or None

                    if not category or amount <= 0:
                        errors.append({"row": i, "error": "成本类别和金额(>0)为必填项"})
                        continue

                    cost_date = None
                    if cost_date_str:
                        try:
                            cost_date = datetime.fromisoformat(cost_date_str)
                        except ValueError:
                            cost_date = datetime.strptime(cost_date_str, "%Y-%m-%d")

                    await self.create_cost({
                        "source_type": source_type,
                        "order_id": str(order_id) if order_id else None,
                        "quote_id": str(quote_id) if quote_id else None,
                        "category": category,
                        "amount": amount,
                        "payment_method": payment_method,
                        "payee_company_name": payee_company_name,
                        "quantity": quantity,
                        "specification": specification,
                        "unit": unit,
                        "unit_price": unit_price,
                        "debt_amount": debt_amount,
                        "description": description,
                        "summary": summary,
                        "cost_date": cost_date.isoformat() if cost_date else None,
                        "remark": remark,
                    }, created_by)
                else:
                    # Standalone import — Excel must include order_no/报价单编号
                    doc_no = str(get("订单编号") or get("报价单编号") or "").strip()
                    if not doc_no:
                        continue
                    category = str(get("成本类别") or "").strip()
                    payment_method = str(get("付款方式") or "").strip() or None
                    payee_company_name = str(get("收款公司") or "").strip() or None
                    quantity_val = get("数量")
                    quantity = float(quantity_val) if quantity_val else None
                    specification = str(get("规格尺寸") or "").strip() or None
                    unit = str(get("单位") or "").strip() or None
                    unit_price_val = get("单价")
                    unit_price = float(unit_price_val) if unit_price_val else None
                    amount_val = get("金额")
                    amount = float(amount_val) if amount_val else 0
                    debt_val = get("欠款金额")
                    debt_amount = float(debt_val) if debt_val else 0
                    cost_date_str = str(get("成本日期") or "").strip() or None
                    description = str(get("说明") or "").strip() or None
                    summary = str(get("成本摘要") or "").strip() or None
                    remark = str(get("备注") or "").strip() or None

                    if not doc_no or not category or amount <= 0:
                        errors.append({"row": i, "error": "单据编号、成本类别和金额(>0)为必填项"})
                        continue

                    # Look up document by doc_no
                    doc_result = await self.db.execute(
                        select(BusinessDocument).where(BusinessDocument.doc_no == doc_no)
                    )
                    doc_obj = doc_result.scalar_one_or_none()
                    if not doc_obj:
                        errors.append({"row": i, "error": f"单据编号「{doc_no}」不存在"})
                        continue

                    cost_date = None
                    if cost_date_str:
                        try:
                            cost_date = datetime.fromisoformat(cost_date_str)
                        except ValueError:
                            cost_date = datetime.strptime(cost_date_str, "%Y-%m-%d")

                    await self.create_cost({
                        "source_type": doc_obj.doc_type,
                        "order_id": str(doc_obj.id),
                        "category": category,
                        "amount": amount,
                        "payment_method": payment_method,
                        "payee_company_name": payee_company_name,
                        "quantity": quantity,
                        "specification": specification,
                        "unit": unit,
                        "unit_price": unit_price,
                        "debt_amount": debt_amount,
                        "description": description,
                        "summary": summary,
                        "cost_date": cost_date.isoformat() if cost_date else None,
                        "remark": remark,
                    }, created_by)
                created += 1
            except Exception as e:
                errors.append({"row": i, "error": str(e)})

        return {"created": created, "errors": errors}

    async def _get_attachment_counts(self, cost_ids: list[UUID]) -> dict[str, int]:
        """Return {cost_id_str: count} for a batch of costs."""
        if not cost_ids:
            return {}
        from sqlalchemy import func
        result = await self.db.execute(
            select(Attachment.related_id, func.count())
            .where(
                Attachment.related_type == "project_cost",
                Attachment.related_id.in_(cost_ids),
            )
            .group_by(Attachment.related_id)
        )
        return {str(row[0]): row[1] for row in result.all()}

    def _to_dict(self, c: ProjectCost) -> dict:
        project_name = None
        doc_no = None
        doc_type = None
        document_id_str = str(c.document_id) if c.document_id else None
        document_item_id_str = str(c.document_item_id) if c.document_item_id else None

        if c.document:
            project_name = c.document.project_name
            doc_no = c.document.doc_no
            doc_type = c.document.doc_type

        # Backward-compat: derive item_name from document_item
        item_name = c.document_item.item_name if c.document_item else None

        return {
            "id": str(c.id),
            "cost_no": c.cost_no,
            # New unified fields
            "document_id": document_id_str,
            "doc_no": doc_no,
            "doc_type": doc_type,
            "document_item_id": document_item_id_str,
            # Backward-compat aliases
            "source_type": doc_type,
            "order_id": document_id_str,
            "quote_id": document_id_str if doc_type == "quote" else None,
            "order_no": doc_no if doc_type == "order" else None,
            "quote_no": doc_no if doc_type == "quote" else None,
            "order_item_id": document_item_id_str if doc_type == "order" else None,
            "quote_item_id": document_item_id_str if doc_type == "quote" else None,
            "group_name": c.group_name,
            "order_item_name": item_name if doc_type == "order" else None,
            "quote_item_name": item_name if doc_type == "quote" else None,
            "document_item_name": item_name,
            "customer_id": str(c.customer_id) if c.customer_id else None,
            "customer_name": c.customer.name if c.customer else None,
            "project_name": project_name,
            "category": c.category,
            "amount": float(c.amount),
            "quantity": float(c.quantity) if c.quantity else None,
            "specification": c.specification,
            "unit": c.unit,
            "unit_price": float(c.unit_price) if c.unit_price else None,
            "payment_method": c.payment_method,
            "payee_company_name": c.payee_company_name,
            "debt_amount": float(c.debt_amount) if c.debt_amount else None,
            "is_debt": c.is_debt,
            "is_settled": c.is_settled,
            "settled_at": c.settled_at.isoformat() if c.settled_at else None,
            "description": c.description,
            "summary": c.summary,
            "cost_date": c.cost_date.isoformat() if c.cost_date else None,
            "receipt_url": c.receipt_url,
            "remark": c.remark,
            "created_by": str(c.created_by) if c.created_by else None,
            "created_at": c.created_at.isoformat() if c.created_at else None,
            "attachment_count": 0,
        }
