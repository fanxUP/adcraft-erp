"""Excel import utility for batch data import."""

from io import BytesIO
from typing import Any

from openpyxl import load_workbook


class ExcelImportResult:
    """Result of an Excel import operation."""

    def __init__(self):
        self.total_rows = 0
        self.succeeded = 0
        self.failed = 0
        self.errors: list[dict] = []  # [{row, message}]

    def to_dict(self) -> dict:
        return {
            "total_rows": self.total_rows,
            "succeeded": self.succeeded,
            "failed": self.failed,
            "errors": self.errors,
        }


def parse_excel(file_bytes: bytes, required_columns: list[str],
                column_map: dict[str, str]) -> tuple[list[dict], list[str]]:
    """Parse an Excel file and return rows as dicts.

    Args:
        file_bytes: Raw Excel file content.
        required_columns: Column header names that must exist in the sheet.
        column_map: Maps Excel header -> internal field name.

    Returns:
        (rows, errors) where rows is a list of dicts mapped to internal names,
        and errors is a list of header-level error messages.
    """
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return [], ["工作簿为空"]

    rows_iter = ws.iter_rows(values_only=True)

    # Read header row
    try:
        headers = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    except StopIteration:
        return [], ["Excel 文件为空"]

    # Validate required columns exist
    errors: list[str] = []
    for col in required_columns:
        if col not in headers:
            errors.append(f"缺少必填列: {col}")
    if errors:
        return [], errors

    # Build index mapping: header index -> internal field name
    field_indices: dict[str, int] = {}
    for i, h in enumerate(headers):
        if h in column_map:
            field_indices[column_map[h]] = i

    rows: list[dict] = []
    for idx, row in enumerate(rows_iter, start=2):
        row_dict: dict[str, Any] = {}
        all_none = True
        for field, col_idx in field_indices.items():
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None:
                all_none = False
            # Convert numeric values cleanly
            if isinstance(val, float):
                # Keep as-is; consumer can handle numeric
                pass
            row_dict[field] = val
        if not all_none:
            row_dict["_excel_row"] = idx
            rows.append(row_dict)

    return rows, []


def format_value(val: Any) -> str | None:
    """Format an Excel cell value to string or None."""
    if val is None:
        return None
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return str(val)
    return str(val).strip() or None


def parse_number(val: Any) -> float | None:
    """Parse an Excel cell value to float or None."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None
