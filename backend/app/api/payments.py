import os
import uuid as _uuid
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, Query, Request, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from io import BytesIO
from uuid import UUID

from app.core.config import settings
from app.core.database import get_db
from app.core.deps import get_current_user
from app.core.permissions import require_role
from app.models.user import User
from app.schemas.payment import PaymentCreate, PaymentVoid, StatementCreate, ExpenseCreate, ExpenseUpdate, ProjectCostCreate, ProjectCostUpdate, DebtSettleCreate
from app.schemas.common import success, success_paginated
from app.services.payment_service import PaymentService, StatementService, ExpenseService
from app.services.project_cost_service import ProjectCostService
from app.services.task_service import AttachmentService
from app.services.operation_log_service import log_operation, OBJ_PAYMENT, OBJ_EXPENSE, OBJ_PROJECT_COST, ACTION_CREATE, ACTION_UPDATE, ACTION_DELETE

pay_router = APIRouter(prefix="/payments", tags=["Payments"])
stmt_router = APIRouter(prefix="/statements", tags=["Statements"])
exp_router = APIRouter(prefix="/expenses", tags=["Expenses"])
cost_router = APIRouter(prefix="/project-costs", tags=["Project Costs"])


# ── Payments ────────────────────────────────────────────────────────────────

@pay_router.get("/")
async def list_payments(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    order_id: str | None = None,
    customer_id: str | None = None,
    is_voided: bool | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    service = PaymentService(db)
    oid = UUID(order_id) if order_id else None
    cid = UUID(customer_id) if customer_id else None
    payments, total = await service.list_payments(page, page_size, oid, cid, is_voided)
    return success_paginated(payments, total, page, page_size)


@pay_router.get("/{payment_id}")
async def get_payment(
    payment_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    service = PaymentService(db)
    payment = await service.get_payment(UUID(payment_id))
    if not payment:
        return {"code": 40401, "message": "收款记录不存在", "data": None}
    return success(payment)


@pay_router.post("/")
async def create_payment(
    data: PaymentCreate,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    service = PaymentService(db)
    payment = await service.create_payment(data.model_dump(), current_user.id)
    await log_operation(db, current_user.id, current_user.real_name or current_user.username,
                        OBJ_PAYMENT, UUID(payment["id"]), ACTION_CREATE,
                        ip_address=request.client.host if request.client else None,
                        after_data={"payment_no": payment["payment_no"], "amount": payment["amount"]})
    return success(payment)


@pay_router.post("/{payment_id}/void")
async def void_payment(
    payment_id: str,
    data: PaymentVoid,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    service = PaymentService(db)
    pid = UUID(payment_id)
    payment = await service.void_payment(pid, data.void_reason)
    await log_operation(db, current_user.id, current_user.real_name or current_user.username,
                        OBJ_PAYMENT, pid, "void",
                        ip_address=request.client.host if request.client else None,
                        after_data={"void_reason": data.void_reason})
    return success(payment)


@pay_router.post("/{payment_id}/upload-receipt")
async def upload_receipt(
    payment_id: str,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    import os, uuid as _uuid
    from datetime import datetime, timezone
    from sqlalchemy import update
    from app.core.config import settings
    from app.models.payment import Payment as PaymentModel

    ext = file.filename.rsplit(".", 1)[-1] if file.filename and "." in file.filename else "bin"
    month_dir = datetime.now(timezone.utc).strftime("%Y%m")
    dest_dir = os.path.join(settings.LOCAL_UPLOAD_DIR, month_dir)
    os.makedirs(dest_dir, exist_ok=True)
    stored_name = f"{_uuid.uuid4()}.{ext}"
    file_path = os.path.join(dest_dir, stored_name)
    contents = await file.read()
    with open(file_path, "wb") as f:
        f.write(contents)
    await file.close()

    service = PaymentService(db)
    p = await service.repo.get_by_id(UUID(payment_id))
    if not p:
        return {"code": 40401, "message": "收款记录不存在", "data": None}
    rel_path = f"{month_dir}/{stored_name}"
    await db.execute(
        update(PaymentModel).where(PaymentModel.id == p.id).values(receipt_url=rel_path)
    )
    await db.flush()
    return success({"file_path": rel_path})


# ── Statements ──────────────────────────────────────────────────────────────

@stmt_router.get("/")
async def list_statements(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    customer_id: str | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    service = StatementService(db)
    cid = UUID(customer_id) if customer_id else None
    stmts, total = await service.list_statements(page, page_size, cid)
    return success_paginated(stmts, total, page, page_size)


@stmt_router.get("/{statement_id}")
async def get_statement(
    statement_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    service = StatementService(db)
    stmt = await service.get_statement(UUID(statement_id))
    if not stmt:
        return {"code": 40401, "message": "对账单不存在", "data": None}
    return success(stmt)


@stmt_router.post("/")
async def create_statement(
    data: StatementCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    service = StatementService(db)
    stmt = await service.create_statement(data.model_dump())
    return success(stmt)


@stmt_router.post("/{statement_id}/confirm")
async def confirm_statement(
    statement_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    service = StatementService(db)
    stmt = await service.confirm_statement(UUID(statement_id), current_user.id)
    return success(stmt)


# ── Expenses ────────────────────────────────────────────────────────────────

@exp_router.get("/")
async def list_expenses(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    category: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    service = ExpenseService(db)
    expenses, total = await service.list_expenses(page, page_size, category, start_date, end_date)
    return success_paginated(expenses, total, page, page_size)


@exp_router.get("/{expense_id}")
async def get_expense(
    expense_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    service = ExpenseService(db)
    expense = await service.get_expense(UUID(expense_id))
    if not expense:
        return {"code": 40401, "message": "支出记录不存在", "data": None}
    return success(expense)


@exp_router.post("/")
async def create_expense(
    data: ExpenseCreate,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    service = ExpenseService(db)
    expense = await service.create_expense(data.model_dump(), current_user.id)
    await log_operation(db, current_user.id, current_user.real_name or current_user.username,
                        OBJ_EXPENSE, UUID(expense["id"]), ACTION_CREATE,
                        ip_address=request.client.host if request.client else None,
                        after_data={"expense_no": expense["expense_no"], "amount": expense["amount"]})
    return success(expense)


@exp_router.put("/{expense_id}")
async def update_expense(
    expense_id: str,
    data: ExpenseUpdate,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    service = ExpenseService(db)
    eid = UUID(expense_id)
    expense = await service.update_expense(eid, data.model_dump(exclude_none=True))
    await log_operation(db, current_user.id, current_user.real_name or current_user.username,
                        OBJ_EXPENSE, eid, ACTION_UPDATE,
                        ip_address=request.client.host if request.client else None)
    return success(expense)


@exp_router.delete("/{expense_id}")
async def delete_expense(
    expense_id: str,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    service = ExpenseService(db)
    eid = UUID(expense_id)
    await service.delete_expense(eid)
    await log_operation(db, current_user.id, current_user.real_name or current_user.username,
                        OBJ_EXPENSE, eid, ACTION_DELETE,
                        ip_address=request.client.host if request.client else None)
    return success(None)


# ── Project Costs ────────────────────────────────────────────────────────────────

@cost_router.get("/")
async def list_project_costs(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    order_id: str | None = None,
    quote_id: str | None = None,
    source_type: str | None = None,
    category: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    service = ProjectCostService(db)
    oid = UUID(order_id) if order_id else None
    qid = UUID(quote_id) if quote_id else None
    costs, total = await service.list_costs(page, page_size, oid, qid, source_type, category, date_from, date_to)
    return success_paginated(costs, total, page, page_size)


@cost_router.get("/summary")
async def get_project_costs_summary(
    order_ids: str = Query(..., description="Comma-separated order UUIDs"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    service = ProjectCostService(db)
    ids = [UUID(oid.strip()) for oid in order_ids.split(",") if oid.strip()]
    costs = await service.get_costs_summary(ids)
    return success({"costs": costs})


@cost_router.get("/template")
async def download_project_cost_template(
    current_user: User = Depends(get_current_user),
):
    """Download an Excel template for importing project costs."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from fastapi.responses import Response

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "成本导入模板"

        header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        headers = ["分项", "成本类别", "付款方式", "收款公司", "数量", "单位", "单价", "金额", "欠款金额", "成本日期", "说明", "成本摘要", "备注"]
        col_widths = [20, 18, 16, 22, 12, 10, 12, 14, 14, 18, 30, 30, 30]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[chr(64 + col_idx)].width = width

        example_data = ["灯箱制作", "人工/工时费", "转账支付", "XX广告制作公司", 2, "套", 250.00, 500.00, 0, "2026-07-08", "安装工人加班费", "灯箱制作成本摘要", "示例数据，可删除"]
        for col_idx, val in enumerate(example_data, 1):
            cell = ws.cell(row=2, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        ws.merge_cells("A4:J4")
        note_cell = ws.cell(row=4, column=1,
            value="说明：成本类别可选值 — 人工/工时费、材料费、租赁费、运输/物流费、安装杂费、办公费、餐费/交通费、差旅费、其他；付款方式可选值 — 现金支付、微信支付、转账支付、对公支付、其它支付")
        note_cell.font = Font(name="微软雅黑", size=9, color="999999", italic=True)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        data = buf.getvalue()

        return Response(
            content=data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=%E9%A1%B9%E7%9B%AE%E6%88%90%E6%9C%AC%E5%AF%BC%E5%85%A5%E6%A8%A1%E6%9D%BF.xlsx",
                "Content-Length": str(len(data)),
            },
        )
    except Exception as e:
        import traceback
        return {"code": 50001, "message": f"生成模板失败: {str(e)}", "detail": traceback.format_exc()}


@cost_router.get("/quotes")
async def list_quotes_for_cost(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    keyword: str | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """获取可登记成本的报价单列表"""
    from app.models.quote import Quote as QuoteModel
    from app.repositories.project_cost_repo import ProjectCostRepository

    from sqlalchemy import or_, select, func

    q = select(QuoteModel).where(QuoteModel.deleted_at.is_(None))
    if keyword:
        fuzzy = f"%{keyword}%"
        q = q.where(
            or_(
                QuoteModel.quote_no.ilike(fuzzy),
                QuoteModel.project_name.ilike(fuzzy),
                QuoteModel.customer_name.ilike(fuzzy),
            )
        )
    count_q = select(func.count()).select_from(q.subquery())
    total = (await db.execute(count_q)).scalar()
    q = q.order_by(QuoteModel.created_at.desc()).offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(q)
    quotes = result.scalars().all()

    # Get cost summaries for these quotes
    repo = ProjectCostRepository(db)
    quote_ids = [q.id for q in quotes]
    cost_map = {}
    if quote_ids:
        cost_map = await repo.get_quote_costs_summary(quote_ids)

    items = []
    for q in quotes:
        items.append({
            "id": str(q.id),
            "quote_no": q.quote_no,
            "project_name": q.project_name,
            "customer_name": q.customer_name,
            "status": q.status,
            "total_amount": float(q.total_amount),
            "cost_amount": float(cost_map.get(str(q.id), 0)),
            "created_at": q.created_at.isoformat() if q.created_at else None,
        })
    return success_paginated(items, total, page, page_size)


@cost_router.get("/debts/list")
async def list_cost_debts(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    keyword: str | None = None,
    is_settled: bool | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """获取成本欠款清单"""
    service = ProjectCostService(db)
    debts, total = await service.list_debts(page, page_size, keyword, is_settled)
    return success_paginated(debts, total, page, page_size)


@cost_router.post("/{cost_id}/settle-debt")
async def settle_cost_debt(
    cost_id: str,
    data: DebtSettleCreate,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """冲红：结算成本欠款"""
    service = ProjectCostService(db)
    try:
        cost = await service.settle_debt(UUID(cost_id), data.model_dump())
    except ValueError as e:
        return {"code": 40001, "message": str(e), "data": None}
    await log_operation(db, current_user.id, current_user.real_name or current_user.username,
                        OBJ_PROJECT_COST, UUID(cost_id), "settle_debt",
                        ip_address=request.client.host if request.client else None,
                        after_data={"settle_amount": data.settle_amount, "payment_method": data.payment_method})
    return success(cost)


@cost_router.get("/{cost_id}")
async def get_project_cost(
    cost_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    service = ProjectCostService(db)
    cost = await service.get_cost(UUID(cost_id))
    if not cost:
        return {"code": 40401, "message": "项目成本记录不存在", "data": None}
    return success(cost)


@cost_router.post("/")
async def create_project_cost(
    data: ProjectCostCreate,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    service = ProjectCostService(db)
    payload = data.model_dump()
    # Validate source_id presence
    source_type = payload.get("source_type", "order")
    if source_type == "quote" and not payload.get("quote_id"):
        return {"code": 40001, "message": "报价单ID不能为空", "data": None}
    if source_type == "order" and not payload.get("order_id"):
        return {"code": 40001, "message": "订单ID不能为空", "data": None}
    try:
        cost = await service.create_cost(payload, current_user.id)
    except ValueError as e:
        return {"code": 40001, "message": str(e), "data": None}
    await log_operation(db, current_user.id, current_user.real_name or current_user.username,
                        OBJ_PROJECT_COST, UUID(cost["id"]), ACTION_CREATE,
                        ip_address=request.client.host if request.client else None,
                        after_data={"cost_no": cost["cost_no"], "amount": cost["amount"], "category": cost["category"]})
    return success(cost)


@cost_router.put("/{cost_id}")
async def update_project_cost(
    cost_id: str,
    data: ProjectCostUpdate,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    service = ProjectCostService(db)
    cid = UUID(cost_id)
    try:
        cost = await service.update_cost(cid, data.model_dump(exclude_none=True))
    except ValueError as e:
        return {"code": 40401, "message": str(e), "data": None}
    await log_operation(db, current_user.id, current_user.real_name or current_user.username,
                        OBJ_PROJECT_COST, cid, ACTION_UPDATE,
                        ip_address=request.client.host if request.client else None)
    return success(cost)


@cost_router.delete("/{cost_id}")
async def delete_project_cost(
    cost_id: str,
    request: Request,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role("admin")),
):
    service = ProjectCostService(db)
    cid = UUID(cost_id)
    try:
        await service.delete_cost(cid)
    except ValueError as e:
        return {"code": 40401, "message": str(e), "data": None}
    await log_operation(db, current_user.id, current_user.real_name or current_user.username,
                        OBJ_PROJECT_COST, cid, ACTION_DELETE,
                        ip_address=request.client.host if request.client else None)
    return success(None)


@cost_router.post("/import")
async def import_project_costs(
    file: UploadFile = File(...),
    order_id: str | None = None,
    quote_id: str | None = None,
    source_type: str = "order",
    request: Request = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        return {"code": 40001, "message": "请上传Excel文件（.xlsx 或 .xls）", "data": None}
    content = await file.read()
    service = ProjectCostService(db)
    try:
        result = await service.import_from_excel(
            BytesIO(content), current_user.id,
            order_id=UUID(order_id) if order_id else None,
            quote_id=UUID(quote_id) if quote_id else None,
            source_type=source_type,
        )
    except Exception as e:
        return {"code": 40001, "message": f"导入失败: {str(e)}", "data": None}
    await log_operation(db, current_user.id, current_user.real_name or current_user.username,
                        OBJ_PROJECT_COST, None, "import",
                        ip_address=request.client.host if request.client else None,
                        after_data={"created": result["created"], "errors": len(result["errors"])})
    return success(result)


# -- Project Cost Attachments (凭证上传) --

@cost_router.get("/{cost_id}/attachments")
async def list_project_cost_attachments(
    cost_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    service = AttachmentService(db)
    atts = await service.list_attachments("project_cost", UUID(cost_id))
    return success(atts)


@cost_router.post("/{cost_id}/upload")
async def upload_project_cost_attachment(
    cost_id: str,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    ext = file.filename.rsplit(".", 1)[-1] if file.filename and "." in file.filename else "bin"
    date_dir = datetime.now(timezone.utc).strftime("%Y%m")
    dest_dir = os.path.join(settings.LOCAL_UPLOAD_DIR, date_dir)
    os.makedirs(dest_dir, exist_ok=True)
    stored_name = f"{_uuid.uuid4().hex}.{ext}"
    file_path = os.path.join(dest_dir, stored_name)
    contents = await file.read()
    with open(file_path, "wb") as f:
        f.write(contents)

    service = AttachmentService(db)
    att = await service.add_attachment(
        related_type="project_cost",
        related_id=UUID(cost_id),
        data={
            "filename": file.filename or stored_name,
            "file_path": f"{date_dir}/{stored_name}",
            "file_size": len(contents),
            "file_type": file.content_type,
        },
        uploaded_by=current_user.id,
    )
    return success(att)


@cost_router.delete("/attachments/{attachment_id}")
async def delete_project_cost_attachment(
    attachment_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    service = AttachmentService(db)
    ok = await service.delete_attachment(UUID(attachment_id))
    if not ok:
        return {"code": 40401, "message": "附件不存在", "data": None}
    return success(None)
