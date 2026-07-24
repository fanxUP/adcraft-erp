"""高空作业车台账模块 — API 路由"""
import json
import uuid as _uuid
from typing import Optional

from fastapi import APIRouter, Body, Depends, HTTPException, Query, Request
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.core.deps import get_current_user
from app.core.permissions import (
    require_permission,
    require_any_role,
    PERM_AERIAL_READ,
    PERM_AERIAL_CREATE,
    PERM_AERIAL_UPDATE,
    PERM_AERIAL_DELETE,
    PERM_AERIAL_FINANCE,
    PERM_AERIAL_WAGE,
    PERM_AERIAL_APPROVE,
)
from app.models.user import User
from app.schemas.aerial import AerialAgentMessageIngest, AerialAgentDraftConfirm, AerialAgentDraftReject
from app.schemas.common import success, success_paginated
from app.services.aerial_service import AerialService


# ── Routers ─────────────────────────────────────────────────────────────────

router = APIRouter(prefix="/aerial/vehicles", tags=["Aerial Vehicles"])
personnel_router = APIRouter(prefix="/aerial/personnel", tags=["Aerial Personnel"])
ledger_router = APIRouter(prefix="/aerial/ledgers", tags=["Aerial Ledgers"])
expense_router = APIRouter(prefix="/aerial/personnel-expenses", tags=["Aerial Personnel Expenses"])
wage_router = APIRouter(prefix="/aerial/personnel-wages", tags=["Aerial Personnel Wages"])
cost_router = APIRouter(prefix="/aerial/vehicle-costs", tags=["Aerial Vehicle Costs"])
safety_router = APIRouter(prefix="/aerial/safety-checks", tags=["Aerial Safety Checks"])
attachment_router = APIRouter(prefix="/aerial/attachments", tags=["Aerial Attachments"])
audit_router = APIRouter(prefix="/aerial/audit-logs", tags=["Aerial Audit Logs"])
dashboard_router = APIRouter(prefix="/aerial/dashboard", tags=["Aerial Dashboard"])
report_router = APIRouter(prefix="/aerial/reports", tags=["Aerial Reports"])


def _svc(db, user, request) -> AerialService:
    ip = request.client.host if request.client else ""
    return AerialService(db, user, ip)


# ── 高空车档案 ──────────────────────────────────────────────────────────────

@router.get("")
async def list_vehicles(
    keyword: str = Query("", description="搜索关键字"),
    status: str = Query("", description="状态筛选"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_READ)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    items, total = await svc.list_vehicles(keyword, status, page, page_size)
    return success_paginated(items, total, page, page_size)


@router.post("")
async def create_vehicle(
    data: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_CREATE)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.create_vehicle(data)
    return success(result)


@router.get("/{vehicle_id}")
async def get_vehicle(
    vehicle_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_READ)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.get_vehicle(vehicle_id)
    return success(result)


@router.patch("/{vehicle_id}")
async def update_vehicle(
    vehicle_id: str,
    data: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_UPDATE)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.update_vehicle(vehicle_id, data)
    return success(result)


# ── 人员 ──────────────────────────────────────────────────────────────────

@personnel_router.get("")
async def list_personnel(
    keyword: str = Query(""),
    status: str = Query(""),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_READ)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    items, total = await svc.list_personnel(keyword, status, page, page_size)
    return success_paginated(items, total, page, page_size)


@personnel_router.post("")
async def create_personnel(
    data: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_CREATE)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.create_personnel(data)
    return success(result)


@personnel_router.get("/{personnel_id}")
async def get_personnel(
    personnel_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_READ)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.get_personnel(personnel_id)
    return success(result)


@personnel_router.patch("/{personnel_id}")
async def update_personnel(
    personnel_id: str,
    data: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_UPDATE)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.update_personnel(personnel_id, data)
    return success(result)


@personnel_router.delete("/{personnel_id}")
async def delete_personnel(
    personnel_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_DELETE)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    try:
        result = await svc.delete_personnel(personnel_id)
        return success(result, message="人员已删除")
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


# ── 每日台账 ────────────────────────────────────────────────────────────────

@ledger_router.get("")
async def list_ledgers(
    date_from: str = Query("", description="开始日期"),
    date_to: str = Query("", description="结束日期"),
    personnel_id: str = Query(""),
    customer_name: str = Query(""),
    work_location: str = Query(""),
    payment_status: str = Query(""),
    audit_status: str = Query(""),
    status: str = Query(""),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_READ)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    items, total = await svc.list_ledgers(
        page=page, page_size=page_size,
        date_from=date_from, date_to=date_to, personnel_id=personnel_id,
        customer_name=customer_name, work_location=work_location,
        payment_status=payment_status, audit_status=audit_status, status=status,
    )
    return success_paginated(items, total, page, page_size)


@ledger_router.post("")
async def create_ledger(
    data: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_CREATE)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.create_ledger(data)
    return success(result)


@ledger_router.get("/{ledger_id}")
async def get_ledger(
    ledger_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_READ)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.get_ledger(ledger_id)
    return success(result)


@ledger_router.patch("/{ledger_id}")
async def update_ledger(
    ledger_id: str,
    data: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_UPDATE)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.update_ledger(ledger_id, data)
    return success(result)


@ledger_router.post("/{ledger_id}/void")
async def void_ledger(
    ledger_id: str,
    data: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_DELETE)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.void_ledger(ledger_id, data.get("reason", ""))
    return success(result)


@ledger_router.post("/{ledger_id}/approve")
async def approve_ledger(
    ledger_id: str,
    data: dict = {},
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_APPROVE)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.approve_ledger(ledger_id, data.get("remark", ""))
    return success(result)


@ledger_router.post("/{ledger_id}/reject")
async def reject_ledger(
    ledger_id: str,
    data: dict = {},
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_APPROVE)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.reject_ledger(ledger_id, data.get("remark", ""))
    return success(result)


# ── 人员垫付 ──────────────────────────────────────────────────────────────

@expense_router.get("")
async def list_expenses(
    date_from: str = Query(""),
    date_to: str = Query(""),
    personnel_id: str = Query(""),
    expense_type: str = Query(""),
    review_status: str = Query(""),
    reimbursement_status: str = Query(""),
    ledger_id: str = Query(""),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_READ)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    items, total = await svc.list_expenses(
        page=page, page_size=page_size,
        date_from=date_from, date_to=date_to, personnel_id=personnel_id,
        expense_type=expense_type, review_status=review_status,
        reimbursement_status=reimbursement_status, ledger_id=ledger_id,
    )
    return success_paginated(items, total, page, page_size)


@expense_router.post("")
async def create_expense(
    data: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_CREATE)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.create_expense(data)
    return success(result)


@expense_router.post("/{expense_id}/review")
async def review_expense(
    expense_id: str,
    data: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_UPDATE)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.review_expense(expense_id, data["status"], data.get("remark", ""))
    return success(result)


@expense_router.post("/{expense_id}/reimburse")
async def reimburse_expense(
    expense_id: str,
    data: dict = {},
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_FINANCE)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.reimburse_expense(expense_id, data.get("remark", ""))
    return success(result)


# ── 人员工资 ──────────────────────────────────────────────────────────────

@wage_router.get("")
async def list_wages(
    wage_month: str = Query(""),
    personnel_id: str = Query(""),
    payment_status: str = Query(""),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_READ)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    items, total = await svc.list_wages(page=page, page_size=page_size, wage_month=wage_month, personnel_id=personnel_id, payment_status=payment_status)
    return success_paginated(items, total, page, page_size)


@wage_router.post("")
async def create_wage(
    data: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_WAGE)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.create_wage(data)
    return success(result)


@wage_router.post("/{wage_id}/pay")
async def pay_wage(
    wage_id: str,
    data: dict = {},
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_FINANCE)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.pay_wage(wage_id, data.get("remark", ""))
    return success(result)


# ── 车辆费用 ────────────────────────────────────────────────────────────────

@cost_router.get("")
async def list_costs(
    date_from: str = Query(""),
    date_to: str = Query(""),
    cost_type: str = Query(""),
    aerial_vehicle_id: str = Query(""),
    review_status: str = Query(""),
    ledger_id: str = Query(""),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_READ)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    items, total = await svc.list_costs(
        page=page, page_size=page_size,
        date_from=date_from, date_to=date_to, cost_type=cost_type,
        aerial_vehicle_id=aerial_vehicle_id, review_status=review_status, ledger_id=ledger_id,
    )
    return success_paginated(items, total, page, page_size)


@cost_router.post("")
async def create_cost(
    data: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_CREATE)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.create_cost(data)
    return success(result)


@cost_router.post("/{cost_id}/review")
async def review_cost(
    cost_id: str,
    data: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_UPDATE)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.review_cost(cost_id, data["status"], data.get("remark", ""))
    return success(result)


# ── 安全检查 ────────────────────────────────────────────────────────────────

@safety_router.get("")
async def list_safety_checks(
    ledger_id: str = Query(..., description="台账ID"),
    check_type: str = Query(""),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_READ)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    items = await svc.list_safety_checks(ledger_id, check_type or None)
    return success(items)


@safety_router.post("")
async def create_safety_check(
    data: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_CREATE)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.create_safety_check(data)
    return success(result)


# ── 附件 ────────────────────────────────────────────────────────────────────

@attachment_router.get("")
async def list_attachments(
    ledger_id: str = Query(..., description="台账ID"),
    attachment_type: str = Query(""),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_READ)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    items = await svc.list_attachments(ledger_id, attachment_type or None)
    return success(items)


@attachment_router.post("")
async def create_attachment(
    data: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_CREATE)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.create_attachment(data)
    return success(result)


@attachment_router.delete("/{attachment_id}")
async def delete_attachment(
    attachment_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_DELETE)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.delete_attachment(attachment_id)
    return success(result)


# ── 审计日志 ────────────────────────────────────────────────────────────────

@audit_router.get("")
async def list_audit_logs(
    ledger_id: str = Query(""),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_READ)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    items, total = await svc.list_audit_logs(ledger_id or None, page, page_size)
    return success_paginated(items, total, page, page_size)


# ── Dashboard ───────────────────────────────────────────────────────────────

@dashboard_router.get("/overview")
async def dashboard_overview(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_READ)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.get_dashboard_overview()
    return success(result)


@dashboard_router.get("/today")
async def dashboard_today(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_READ)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.get_today_ledgers()
    return success(result)


@dashboard_router.get("/reminders")
async def dashboard_reminders(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_READ)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.get_reminders()
    return success(result)


# ── 报表 ────────────────────────────────────────────────────────────────────

@report_router.get("/daily")
async def report_daily(
    date: str = Query(..., description="日期 YYYY-MM-DD"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_READ)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.get_report_daily(date)
    return success(result)


@report_router.get("/monthly")
async def report_monthly(
    month: str = Query(..., description="月份 YYYY-MM"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_READ)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.get_report_monthly(month)
    return success(result)


@report_router.get("/receivables")
async def report_receivables(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_READ)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.get_report_receivables(page, page_size)
    return success(result)


@report_router.get("/reimbursements")
async def report_reimbursements(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_READ)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.get_report_reimbursements(page, page_size)
    return success(result)


@report_router.get("/costs")
async def report_costs(
    month: str = Query(""),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_READ)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.get_report_costs(month or None)
    return success(result)


@report_router.get("/personnel-summary")
async def report_personnel_summary(
    month: str = Query(..., description="月份 YYYY-MM"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_READ)),
    request: Request = None,
):
    svc = _svc(db, current_user, request)
    result = await svc.get_report_personnel_summary(month)
    return success(result)


@report_router.get("/export/ledgers")
async def export_ledgers(
    start_date: str = Query(..., description="开始日期 YYYY-MM-DD"),
    end_date: str = Query(..., description="结束日期 YYYY-MM-DD"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_READ)),
    request: Request = None,
):
    """导出台账为 Excel"""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    import io

    svc = _svc(db, current_user, request)
    items, _ = await svc.list_ledgers(page=1, page_size=10000, start_date=start_date, end_date=end_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "出车台账"

    headers = ["台账编号", "日期", "车牌号", "人员", "客户", "作业地点", "作业类型",
               "应收", "实收", "未收", "工资", "报销", "毛利", "状态", "收款状态"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for item in items:
        ws.append([
            item.get("ledger_no"), item.get("work_date"), item.get("plate_number"),
            item.get("name"), item.get("customer_name"), item.get("work_location"),
            item.get("work_type"), item.get("receivable_amount"), item.get("received_amount"),
            item.get("unpaid_amount"), item.get("personnel_wage_amount"),
            item.get("reimbursement_amount"), item.get("gross_profit"),
            item.get("status"), item.get("payment_status"),
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"出车台账_{start_date}_{end_date}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


@report_router.get("/export/wages")
async def export_wages(
    month: str = Query(..., description="月份 YYYY-MM"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_READ)),
    request: Request = None,
):
    """导出工资为 Excel"""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font
    import io

    svc = _svc(db, current_user, request)
    items, _ = await svc.list_wages(page=1, page_size=10000, wage_month=month)

    wb = Workbook()
    ws = wb.active
    ws.title = "人员工资"

    headers = ["工资月份", "人员", "工资类型", "基本工资", "出车工资", "计时工资",
               "提成", "补贴", "扣款", "实发工资", "发放状态"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for item in items:
        ws.append([
            item.get("wage_month"), item.get("name"), item.get("wage_type"),
            item.get("base_wage"), item.get("trip_wage"), item.get("hourly_wage"),
            item.get("commission_amount"), item.get("allowance_amount"),
            item.get("deduction_amount"), item.get("final_wage_amount"),
            item.get("payment_status"),
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"人员工资_{month}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


# ── Agent 草稿路由 ──────────────────────────────────────────────────────────

agent_router = APIRouter(prefix="/aerial/agent", tags=["高空车Agent"])


def _agent_svc(db, current_user, request):
    from app.services.aerial_agent_service import AerialAgentService
    ip = request.client.host if request and request.client else ""
    return AerialAgentService(db, current_user, ip)


@agent_router.post("/messages/ingest")
async def agent_ingest_message(
    data: AerialAgentMessageIngest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_CREATE)),
    request: Request = None,
):
    """接收 Agent 消息，识别意图并生成草稿"""
    svc = _agent_svc(db, current_user, request)
    result = await svc.ingest_message(data.model_dump())
    return success(result)


@agent_router.get("/drafts")
async def agent_list_drafts(
    status: Optional[str] = Query(None, description="草稿状态: pending/confirmed/rejected/expired"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_READ)),
    request: Request = None,
):
    """列出 Agent 草稿"""
    svc = _agent_svc(db, current_user, request)
    skip = (page - 1) * page_size
    rows, total = await svc.list_drafts(status=status, skip=skip, limit=page_size)
    items = []
    for r in rows:
        extracted = {}
        if r.extracted_json:
            try:
                extracted = json.loads(r.extracted_json)
            except Exception:
                pass
        items.append({
            "id": str(r.id),
            "platform": r.platform,
            "conversation_id": r.conversation_id,
            "sender_id": r.sender_id,
            "sender_name": r.sender_name,
            "raw_message": r.raw_message,
            "intent": r.intent,
            "confidence": r.confidence,
            "risk_level": r.risk_level,
            "extracted": extracted,
            "suggested_action": r.suggested_action,
            "status": r.status,
            "confirmed_by": str(r.confirmed_by) if r.confirmed_by else None,
            "confirmed_at": r.confirmed_at.isoformat() if r.confirmed_at else None,
            "reject_reason": r.reject_reason,
            "created_ledger_id": str(r.created_ledger_id) if r.created_ledger_id else None,
            "created_expense_id": str(r.created_expense_id) if r.created_expense_id else None,
            "created_cost_id": str(r.created_cost_id) if r.created_cost_id else None,
            "created_at": r.created_at.isoformat() if r.created_at else "",
        })
    return success_paginated(items, total, page, page_size)


@agent_router.get("/drafts/{draft_id}")
async def agent_get_draft(
    draft_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_READ)),
    request: Request = None,
):
    """获取草稿详情"""
    svc = _agent_svc(db, current_user, request)
    r = await svc.get_draft(_uuid.UUID(draft_id))
    if not r:
        raise HTTPException(status_code=404, detail="草稿不存在")
    extracted = {}
    if r.extracted_json:
        try:
            extracted = json.loads(r.extracted_json)
        except Exception:
            pass
    return success({
        "id": str(r.id),
        "platform": r.platform,
        "conversation_id": r.conversation_id,
        "sender_id": r.sender_id,
        "sender_name": r.sender_name,
        "raw_message": r.raw_message,
        "intent": r.intent,
        "confidence": r.confidence,
        "risk_level": r.risk_level,
        "extracted": extracted,
        "suggested_action": r.suggested_action,
        "status": r.status,
        "confirmed_by": str(r.confirmed_by) if r.confirmed_by else None,
        "confirmed_at": r.confirmed_at.isoformat() if r.confirmed_at else None,
        "reject_reason": r.reject_reason,
        "created_ledger_id": str(r.created_ledger_id) if r.created_ledger_id else None,
        "created_expense_id": str(r.created_expense_id) if r.created_expense_id else None,
        "created_cost_id": str(r.created_cost_id) if r.created_cost_id else None,
        "created_at": r.created_at.isoformat() if r.created_at else "",
    })


@agent_router.post("/drafts/{draft_id}/confirm")
async def agent_confirm_draft(
    draft_id: str,
    data: AerialAgentDraftConfirm = Body(AerialAgentDraftConfirm()),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_UPDATE)),
    request: Request = None,
):
    """确认草稿，写入正式台账"""
    svc = _agent_svc(db, current_user, request)
    result = await svc.confirm_draft(_uuid.UUID(draft_id), data.adjustments)
    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("error", "确认失败"))
    return success(result)


@agent_router.post("/drafts/{draft_id}/reject")
async def agent_reject_draft(
    draft_id: str,
    data: AerialAgentDraftReject = Body(AerialAgentDraftReject()),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_permission(PERM_AERIAL_UPDATE)),
    request: Request = None,
):
    """拒绝草稿"""
    svc = _agent_svc(db, current_user, request)
    result = await svc.reject_draft(_uuid.UUID(draft_id), data.reason)
    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("error", "拒绝失败"))
    return success(result)
